"""
TASK-040 成绩导出 单元测试
"""

from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import Group
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.accounts.roles import add_user_role, ROLE_CHOICES
from apps.organizations.models import Class, College
from apps.evaluations.models import EvaluationBatch
from apps.statistics.models import ScoreSummary
from apps.statistics.views import ExportViewSet


class ExportTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        for role_name in ROLE_CHOICES:
            Group.objects.get_or_create(name=role_name)

        cls.college = College.objects.create(code="CS", name="CS")
        cls.class_obj = Class.objects.create(college=cls.college, name="c1", grade="2025")
        cls.other_class = Class.objects.create(college=cls.college, name="c2", grade="2025")

        cls.super_admin = User.objects.create_user(username="sa", password="p", real_name="超管")
        add_user_role(cls.super_admin, "super_admin")

        cls.counselor = User.objects.create_user(username="tc", password="p", real_name="辅导员")
        add_user_role(cls.counselor, "counselor")
        cls.class_obj.counselor = cls.counselor; cls.class_obj.save()

        cls.student = User.objects.create_user(username="s1", password="p", real_name="张三",
                                                student_class=cls.class_obj, student_no="001")
        add_user_role(cls.student, "student")

        cls.batch = EvaluationBatch.objects.create(
            name="测试批次", start_time=timezone.now(),
            end_time=timezone.now() + timezone.timedelta(days=30),
            status=EvaluationBatch.Status.RUNNING,
        )

        ScoreSummary.objects.create(
            student=cls.student, batch=cls.batch,
            total_score=Decimal("10.0"), ranking=1,
        )
        # unranked
        u2 = User.objects.create_user(username="s2", password="p", real_name="李四",
                                       student_class=cls.class_obj, student_no="002")
        add_user_role(u2, "student")
        ScoreSummary.objects.create(student=u2, batch=cls.batch,
                                     total_score=Decimal("5.0"), ranking=0)

        cls.factory = APIRequestFactory()

    def _export(self, user, params=""):
        request = self.factory.get(f"/api/v1/export/scores/{params}")
        force_authenticate(request, user=user)
        view = ExportViewSet.as_view({"get": "export_scores"})
        return view(request)

    def _parse(self, response):
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = [[cell.value for cell in row] for row in ws.iter_rows()]
        return rows

    # ============================================================
    def test_01_content_type(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        self.assertEqual(response.status_code, 200)
        self.assertIn("openxmlformats", response["Content-Type"])

    def test_02_content_disposition(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        cd = response["Content-Disposition"]
        # Django encodes non-ASCII filenames via RFC 2047
        self.assertTrue(cd.startswith("=?utf-8?b?"))

    def test_03_batch_required(self):
        response = self._export(self.super_admin, "")
        self.assertEqual(response.status_code, 400)

    def test_04_empty_data_still_generates_xlsx(self):
        empty_batch = EvaluationBatch.objects.create(
            name="空批次", start_time=timezone.now(),
            end_time=timezone.now() + timezone.timedelta(days=1),
            status=EvaluationBatch.Status.RUNNING,
        )
        response = self._export(self.super_admin, "?batch=" + str(empty_batch.id))
        self.assertEqual(response.status_code, 200)

    def test_05_excludes_ranking_zero(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        rows = self._parse(response)
        ranking_values = [r[0] for r in rows[1:]]  # skip header
        self.assertNotIn(0, ranking_values)

    def test_06_ranking_order(self):
        # s2 already has ranking=0, update it to 2 to test sort
        u2 = User.objects.get(username="s2")
        ScoreSummary.objects.filter(student=u2, batch=self.batch).update(ranking=2)
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        rows = self._parse(response)
        rankings = [r[0] for r in rows[1:]]
        self.assertEqual(rankings, sorted(rankings))

    def test_07_header_bold_and_frozen(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertTrue(ws.cell(1, 1).font.bold)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_08_counselor_other_class_forbidden(self):
        response = self._export(
            self.counselor,
            f"?batch={self.batch.id}&class_id={self.other_class.id}"
        )
        self.assertEqual(response.status_code, 404)
