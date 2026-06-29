"""
TASK-041 成果导出 单元测试
"""

from io import BytesIO

from django.contrib.auth.models import Group
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate
from django.utils import timezone
from openpyxl import load_workbook

from apps.accounts.models import User
from apps.accounts.roles import add_user_role, ROLE_CHOICES
from apps.organizations.models import Class, College
from apps.evaluations.models import EvaluationBatch
from apps.achievements.models import Achievement, AchievementCategory
from apps.statistics.views import ExportViewSet


class ExportAchievementsTest(TestCase):
    @classmethod
    def setUpTestData(cls):
        for role_name in ROLE_CHOICES:
            Group.objects.get_or_create(name=role_name)

        cls.college = College.objects.create(code="CS", name="CS")
        cls.class_obj = Class.objects.create(college=cls.college, name="c1", grade="2025")
        cls.other_class = Class.objects.create(college=cls.college, name="c2", grade="2025")

        cls.super_admin = User.objects.create_user(username="sa", password="p", real_name="超管")
        add_user_role(cls.super_admin, "super_admin")

        cls.counselor = User.objects.create_user(username="tc", password="p", real_name="辅导员")
        add_user_role(cls.counselor, "counselor")
        cls.class_obj.counselor = cls.counselor; cls.class_obj.save()

        cls.student = User.objects.create_user(username="s1", password="p", real_name="张三",
                                                student_class=cls.class_obj, student_no="001")
        add_user_role(cls.student, "student")

        cls.batch = EvaluationBatch.objects.create(
            name="测试批次", start_time=timezone.now(),
            end_time=timezone.now() + timezone.timedelta(days=30),
            status=EvaluationBatch.Status.RUNNING,
        )
        cls.category = AchievementCategory.objects.create(name="学科竞赛")

        Achievement.objects.create(
            student=cls.student, class_obj=cls.class_obj,
            batch=cls.batch, category=cls.category,
            title="成果1", status="approved", score=8.0,
        )

        cls.factory = APIRequestFactory()

    def _export(self, user, params=""):
        request = self.factory.get(f"/api/v1/export/achievements/{params}")
        force_authenticate(request, user=user)
        view = ExportViewSet.as_view({"get": "export_achievements"})
        return view(request)

    # ============================================================
    def test_01_content_type(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        self.assertEqual(response.status_code, 200)
        self.assertIn("openxmlformats", response["Content-Type"])

    def test_02_filename(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        self.assertTrue(response["Content-Disposition"].startswith("=?utf-8?b?"))

    def test_03_batch_required(self):
        response = self._export(self.super_admin, "")
        self.assertEqual(response.status_code, 400)

    def test_04_invalid_status(self):
        response = self._export(self.super_admin, f"?batch={self.batch.id}&status=xxx")
        self.assertEqual(response.status_code, 400)

    def test_05_default_status_approved(self):
        Achievement.objects.create(
            student=self.student, class_obj=self.class_obj,
            batch=self.batch, category=self.category,
            title="draft成果", status="draft",
        )
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)  # only approved

    def test_06_header_frozen(self):
        response = self._export(self.super_admin, "?batch=" + str(self.batch.id))
        wb = load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertTrue(ws.cell(1, 1).font.bold)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_07_counselor_only_managed(self):
        response = self._export(self.counselor, f"?batch={self.batch.id}")
        self.assertEqual(response.status_code, 200)

    def test_08_student_forbidden(self):
        response = self._export(self.student, "?batch=" + str(self.batch.id))
        self.assertEqual(response.status_code, 403)
