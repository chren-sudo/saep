"""
statistics 视图
"""

from django.http import Http404, HttpResponse
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.mixins import ListModelMixin
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet

from apps.achievements.models import Achievement
from apps.achievements.views import _get_achievement_queryset

from .models import ScoreSummary
from .serializers import ScoreSummarySerializer


# ============================================================
# 共享查询 — ScoreViewSet.ranking + ExportViewSet 共用
# ============================================================

def _get_score_queryset(user, batch_id, class_id=None):
    """按权限获取已排名的 ScoreSummary QuerySet"""
    group_names = set(user.groups.values_list("name", flat=True))

    qs = ScoreSummary.objects.filter(
        batch_id=int(batch_id), ranking__gt=0
    ).select_related("student__student_class", "batch")

    # counselor: 仅管理的班级
    if "counselor" in group_names:
        if not class_id:
            raise ValidationError("辅导员请指定 class_id 参数")
        if not user.managed_classes.filter(id=class_id).exists():
            raise Http404
        qs = qs.filter(student__student_class_id=int(class_id))

    # college_admin / super_admin: 可选 class_id
    elif "college_admin" in group_names or "super_admin" in group_names:
        if class_id:
            qs = qs.filter(student__student_class_id=int(class_id))

    # student / evaluator: 只看自己的
    else:
        qs = qs.filter(student=user)

    return qs.order_by("ranking", "student__student_no")


# ============================================================
# ScoreViewSet (TASK-035/036)
# ============================================================

class ScoreViewSet(GenericViewSet, ListModelMixin):
    """成绩查询"""

    permission_classes = [IsAuthenticated]
    serializer_class = ScoreSummarySerializer
    filterset_fields = ["batch"]

    def get_queryset(self):
        return ScoreSummary.objects.filter(
            student=self.request.user
        ).select_related("batch", "student__student_class")

    @action(detail=False, methods=["get"], url_path="ranking")
    def ranking(self, request):
        batch_id = request.query_params.get("batch")
        if not batch_id:
            raise ValidationError("batch 参数为必填项")

        qs = _get_score_queryset(
            request.user, batch_id,
            request.query_params.get("class_id"),
        )

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)

        serializer = self.get_serializer(qs, many=True)
        return Response({"code": 0, "message": "success", "data": serializer.data})


# ============================================================
# ExportViewSet (TASK-040)
# ============================================================

class ExportViewSet(GenericViewSet):
    """数据导出"""

    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        user = self.request.user
        if not user.is_authenticated:
            return [IsAuthenticated()]
        # 仅 counselor / college_admin / super_admin 可导出
        if not user.groups.filter(name__in=["counselor", "college_admin", "super_admin"]).exists():
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("无权导出数据")
        return [IsAuthenticated()]

    @action(detail=False, methods=["get"], url_path="scores")
    def export_scores(self, request):
        """GET /export/scores/?batch=3&class_id=3"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        batch_id = request.query_params.get("batch")
        if not batch_id:
            raise ValidationError("batch 参数为必填项")

        qs = _get_score_queryset(
            request.user, batch_id,
            request.query_params.get("class_id"),
        )

        batch_name = qs[0].batch.name if qs.exists() else f"批次{batch_id}"

        wb = Workbook()
        ws = wb.active
        ws.title = "成绩排名"

        # 表头
        headers = ["排名", "学号", "姓名", "班级", "总分"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        # 数据行
        for row_idx, s in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=s.ranking)
            ws.cell(row=row_idx, column=2, value=s.student.student_no)
            ws.cell(row=row_idx, column=3, value=s.student.real_name)
            ws.cell(row=row_idx, column=4, value=s.student.student_class.name)
            ws.cell(row=row_idx, column=5, value=float(s.total_score))

        # 自动列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        # 冻结首行
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{batch_name}_成绩排名.xlsx"'
        )

        # TASK-045: 导出日志
        from apps.system.services import write_operation_log
        from apps.system.models import OperationLog
        class_id = request.query_params.get("class_id", "ALL")
        write_operation_log(request.user, request,
            module=OperationLog.Module.EXPORT,
            action=OperationLog.Action.EXPORT,
            detail=f"导出成绩排名 batch={batch_id}, class={class_id}, type=scores",
        )

        return response

    @action(detail=False, methods=["get"], url_path="achievements")
    def export_achievements(self, request):
        """GET /export/achievements/?batch=3&status=approved"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        batch_id = request.query_params.get("batch")
        if not batch_id:
            raise ValidationError("batch 参数为必填项")

        status = request.query_params.get("status", "approved")
        valid_statuses = set(Achievement.Status.values)
        if status not in valid_statuses:
            raise ValidationError(f"无效的状态值: {status}，可选: {', '.join(sorted(valid_statuses))}")

        qs = _get_achievement_queryset(request.user)
        qs = qs.filter(batch_id=int(batch_id), status=status).order_by(
            "class_obj_id", "student__student_no", "title"
        )

        batch_name = qs[0].batch.name if qs.exists() else f"批次{batch_id}"

        wb = Workbook()
        ws = wb.active
        ws.title = "成果列表"

        headers = ["学号", "姓名", "班级", "成果名称", "分类", "等级", "得分", "状态", "获得日期", "提交时间"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for row_idx, a in enumerate(qs, 2):
            ws.cell(row=row_idx, column=1, value=a.student.student_no)
            ws.cell(row=row_idx, column=2, value=a.student.real_name)
            ws.cell(row=row_idx, column=3, value=a.class_obj.name)
            ws.cell(row=row_idx, column=4, value=a.title)
            ws.cell(row=row_idx, column=5, value=a.category.name)
            ws.cell(row=row_idx, column=6, value=a.level)
            ws.cell(row=row_idx, column=7, value=float(a.score) if a.score else None)
            ws.cell(row=row_idx, column=8, value=a.get_status_display())
            ws.cell(row=row_idx, column=9, value=str(a.achievement_date) if a.achievement_date else "")
            ws.cell(row=row_idx, column=10, value=str(a.submitted_at) if a.submitted_at else "")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 16

        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{batch_name}_成果列表.xlsx"'
        )

        # TASK-045: 导出日志
        from apps.system.services import write_operation_log
        from apps.system.models import OperationLog
        class_id = request.query_params.get("class_id", "ALL")
        status_val = request.query_params.get("status", "approved")
        write_operation_log(request.user, request,
            module=OperationLog.Module.EXPORT,
            action=OperationLog.Action.EXPORT,
            detail=f"导出成果列表 batch={batch_id}, class={class_id}, status={status_val}, type=achievements",
        )

        return response
