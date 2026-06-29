"""
accounts 视图
"""

import io

from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework_simplejwt.exceptions import TokenError
from rest_framework_simplejwt.tokens import RefreshToken

from .permissions import IsCollegeAdminOrSuperAdmin
from .serializers import LoginSerializer, ProfileSerializer, StudentImportSerializer, TokenRefreshSerializer


@api_view(["GET"])
@permission_classes([AllowAny])
def health_check(request):
    """健康检查接口"""
    return Response(
        {
            "code": 0,
            "message": "success",
            "data": {
                "status": "ok",
            },
        }
    )


@api_view(["POST"])
@permission_classes([AllowAny])
def login(request):
    """登录

    POST /api/v1/auth/login/
    """
    serializer = LoginSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    user = serializer.validated_data["user"]
    refresh = RefreshToken.for_user(user)

    # TASK-043: 登录日志
    from apps.system.services import write_operation_log
    from apps.system.models import OperationLog
    write_operation_log(user, request, OperationLog.Module.AUTH, OperationLog.Action.LOGIN)

    return Response(
        {
            "code": 0,
            "message": "success",
            "data": {
                "access": str(refresh.access_token),
                "refresh": str(refresh),
            },
        }
    )


@api_view(["POST"])
@permission_classes([AllowAny])
def token_refresh(request):
    """刷新 Token

    POST /api/v1/auth/refresh/
    """
    serializer = TokenRefreshSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    try:
        refresh = RefreshToken(serializer.validated_data["refresh"])
        return Response(
            {
                "code": 0,
                "message": "success",
                "data": {
                    "access": str(refresh.access_token),
                },
            }
        )
    except TokenError:
        return Response(
            {
                "code": 401,
                "message": "Token 无效或已过期",
                "data": None,
            },
            status=status.HTTP_401_UNAUTHORIZED,
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def profile(request):
    """获取当前用户信息

    GET /api/v1/auth/profile/
    """
    serializer = ProfileSerializer(request.user)
    return Response(
        {
            "code": 0,
            "message": "success",
            "data": serializer.data,
        }
    )


# ============================================================
# Student Import (TASK-007)
# ============================================================

IMPORT_TEMPLATE_HEADERS = ["学号", "姓名", "手机号", "学院代码", "班级名称"]


@api_view(["GET"])
@permission_classes([IsCollegeAdminOrSuperAdmin])
def student_template(request):
    """下载学生导入 Excel 模板

    GET /api/v1/students/template/
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(IMPORT_TEMPLATE_HEADERS)

    # 添加示例行
    ws.append(["20250001", "张三", "13800000001", "CS", "计算机科学与技术1班"])
    ws.append(["20250002", "李四", "13800000002", "SE", "软件工程1班"])

    # 设置列宽
    col_widths = [14, 10, 16, 12, 24]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="student_import_template.xlsx"'
    return response


@api_view(["POST"])
@permission_classes([IsCollegeAdminOrSuperAdmin])
def student_import(request):
    """批量导入学生

    POST /api/v1/students/import/
    """
    serializer = StudentImportSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)

    file = serializer.validated_data["file"]
    result = _process_import(file)

    return Response(
        {
            "code": 0,
            "message": "success",
            "data": result,
        }
    )


def _process_import(file):
    """处理 Excel 导入的核心逻辑"""
    from django.contrib.auth.models import Group

    from .models import User
    from apps.organizations.models import Class, College

    wb = load_workbook(file, read_only=True)
    ws = wb.active

    total = 0
    success = 0
    errors = []
    seen_student_nos = set()

    student_group = Group.objects.get(name="student")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        student_no = _cell_str(row, 0)
        real_name = _cell_str(row, 1)
        phone = _cell_str(row, 2)
        college_code = _cell_str(row, 3)
        class_name = _cell_str(row, 4)

        total += 1

        # 必填校验
        if not student_no:
            errors.append({"row": row_idx, "student_no": "", "detail": "学号为空"})
            continue
        if not real_name:
            errors.append({"row": row_idx, "student_no": student_no, "detail": "姓名为空"})
            continue
        if not college_code:
            errors.append({"row": row_idx, "student_no": student_no, "detail": "学院代码为空"})
            continue
        if not class_name:
            errors.append({"row": row_idx, "student_no": student_no, "detail": "班级名称为空"})
            continue

        # Excel 内部学号重复
        if student_no in seen_student_nos:
            errors.append({"row": row_idx, "student_no": student_no, "detail": "Excel 内学号重复"})
            continue
        seen_student_nos.add(student_no)

        # 学号已存在
        if User.objects.filter(username=student_no).exists():
            errors.append({"row": row_idx, "student_no": student_no, "detail": "学号已存在，跳过"})
            continue

        # 学院校验
        try:
            college = College.objects.get(code=college_code.upper())
        except College.DoesNotExist:
            errors.append({"row": row_idx, "student_no": student_no, "detail": f"学院代码 '{college_code}' 不存在"})
            continue

        # 班级校验（限定学院）
        try:
            student_class = Class.objects.get(college=college, name=class_name)
        except Class.DoesNotExist:
            errors.append({"row": row_idx, "student_no": student_no, "detail": f"班级 '{class_name}' 不在学院 '{college_code}' 下"})
            continue

        # 创建用户（默认密码 = 学号后6位）
        user = User.objects.create_user(
            username=student_no,
            password=student_no[-6:],
            real_name=real_name,
            student_no=student_no,
            phone=phone,
            student_class=student_class,
        )
        user.groups.add(student_group)
        success += 1

    wb.close()

    return {
        "total": total,
        "success": success,
        "skipped": len(errors),
        "errors": errors,
    }


def _cell_str(row, index):
    """从 openpyxl 行元组中安全取字符串值"""
    if index >= len(row):
        return ""
    val = row[index]
    if val is None:
        return ""
    return str(val).strip()

